import os
from openpyxl import load_workbook

# ============================================================================
# PEMETAAN LENGKAP KOLOM SHEET "Inject" (1-75)
# ============================================================================
# 01 (A ): NO
# 02 (B ): Employee ID *                  <- [DIISI: id]
# 03 (C ): Barcode
# 04 (D ): First Name *                   <- [DIISI: nama_depan]
# 05 (E ): Last Name                      <- [DIISI: nama_belakang]
# 06 (F ): Email *
# 07 (G ): NIK (NPWP 16 digit)            <- [DIISI: nik_ktp]
# 08 (H ): Citizen ID Address             <- [DIISI: alamat]
# 09 (I ): Residential Address
# 10 (J ): Place of Birth                 <- [DIISI: tempat_lahir]
# 11 (K ): Date of Birth *                <- [DIISI: tanggal_lahir]
# 12 (L ): Phone Number
# 13 (M ): Additional Phone Number
# 14 (N ): Gender *                       <- [DIISI: jenis_kelamin]
# 15 (O ): Marital Status *               <- [DIISI: status_perkawinan]
# 16 (P ): Religion *                     <- [DIISI: agama]
# 17 (Q ): Organization Name *
# 18 (R ): Job Position *
# 19 (S ): Job Level *
# 20 (T ): Grade
# 21 (U ): Class
# 22 (V ): Employment Status *
# 23 (W ): Join Date *
# 24 (X ): End Employment Status Date
# 25 (Y ): Sign Date
# 26 (Z ): NPWP                           <- [DIISI: npwp]
# 27 (AA): Taxable Date
# 28 (AB): PTKP Status *                  <- [DIISI: status_ptkp]
# 29 (AC): Bank Name
# 30 (AD): Bank Account                   <- [DIISI: rekening_bank]
# 31 (AE): Bank Account Holder
# 32 (AF): BPJS Ketenagakerjaan           <- [DIISI: bpjs_tk]
# 33 (AG): BPJS Kesehatan                 <- [DIISI: bpjs_kes]
# 34 (AH): Resign Date
# 35 (AI): Branch Name
# 36 (AJ): Type Salary
# 37 (AK): Overtime Status
# 38 (AL): Passport
# 39 (AM): Passport Expired Date
# 40 (AN): Blood Type                     <- [DIISI: golongan_darah]
# 41 (AO): Postal Code
# 42 (AP): BPJS Kesehatan Family
# 43 (AQ): Employee Tax Status*
# 44 (AR): JHT Config
# 45 (AS): Tax Config
# 46 (AT): BPJS Kesehatan Config
# 47 (AU): Jaminan Pensiun Config
# 48 (AV): NPP BPJS Ketenagakerjaan
# 49 (AW): Beginning Netto
# 50 (AX): PPH 21 Paid
# 51 (AY): Ekspatriat DN Date
# 52 (AZ): Nationality Code
# 53 (BA): BPJS Ketenagakerjaan Date
# 54 (BB): BPJS Kesehatan Date
# 55 (BC): Jaminan Pensiun Date
# 56 (BD): Payment Schedule
# 57 (BE): Salary Config
# 58 (BF): Currency
# 59 (BG): Cost Center
# 60 (BH): Cost Center Category
# 61 (BI): Work Schedule
# 62 (BJ): Overtime Working Day Default
# 63 (BK): Overtime Day Off Default
# 64 (BL): Overtime National Holiday Default
# 65 (BM): Split Payment Policy
# 66 (BN): Bank Name Secondary
# 67 (BO): Bank Account  Secondary
# 68 (BP): Bank Account Holder  Secondary
# 69 (BQ): Bank Name Tertiary
# 70 (BR): Bank Account Tertiary
# 71 (BS): Bank Account Holder Tertiary
# 72 (BT): Jenis Dok. Referensi Bukti Potong
# 73 (BU): Nomor Dok. Referensi Bukti Potong
# 74 (BV): Tanggal Dok. Referensi Bukti Potong
# 75 (BW): TIN (Taxpayer Identification Number)
# 76 (BX): [KOLOM EKSTRA: Catatan Dokumen/Error Lapis 1 & 2] <- [DIISI: catatan_dokumen]
# ============================================================================

def append_to_excel(excel_path: str, employee_meta: dict, extracted_data: dict) -> None:
    """
    Menyisipkan (append) data karyawan ke baris kosong pertama secara eksplisit per sel 
    berdasarkan index kolom presisi (1-based index) dari HRD.
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"File Excel tidak ditemukan di: {excel_path}")
        
    try:
        workbook = load_workbook(excel_path)
        # Menargetkan sheet secara spesifik, sesuai arahan (Sheet: 'Inject')
        if "Inject" in workbook.sheetnames:
            sheet = workbook["Inject"]
        else:
            sheet = workbook.active
        
        next_row = sheet.max_row + 1
        
        # Ekstraksi Nama: Prioritaskan hasil parsing AI Pydantic, 
        # namun fallback ke nama dari folder jika AI gagal mengisi.
        nama_depan = extracted_data.get("nama_depan")
        nama_belakang = extracted_data.get("nama_belakang")
        
        if not nama_depan and not nama_belakang:
            full_name = employee_meta.get("nama", "")
            parts = full_name.split(" ", 1)
            nama_depan = parts[0] if len(parts) > 0 else ""
            nama_belakang = parts[1] if len(parts) > 1 else ""

        # --- Explicit Column Mapping (Sesuai Referensi Di Atas) ---
        # Menginisialisasi/menulis SEMUA 75 kolom tanpa terkecuali agar bersih dan terstruktur.
        sheet.cell(row=next_row, column=1).value = ""  # NO
        sheet.cell(row=next_row, column=2).value = employee_meta.get("id", "")  # Employee ID *
        sheet.cell(row=next_row, column=3).value = ""  # Barcode
        sheet.cell(row=next_row, column=4).value = nama_depan or ""  # First Name *
        sheet.cell(row=next_row, column=5).value = nama_belakang or ""  # Last Name
        sheet.cell(row=next_row, column=6).value = ""  # Email *
        sheet.cell(row=next_row, column=7).value = extracted_data.get("nik_ktp") or ""  # NIK (NPWP 16 digit)
        sheet.cell(row=next_row, column=8).value = extracted_data.get("alamat") or ""  # Citizen ID Address
        sheet.cell(row=next_row, column=9).value = ""  # Residential Address
        sheet.cell(row=next_row, column=10).value = extracted_data.get("tempat_lahir") or ""  # Place of Birth
        sheet.cell(row=next_row, column=11).value = extracted_data.get("tanggal_lahir") or ""  # Date of Birth *
        sheet.cell(row=next_row, column=12).value = ""  # Phone Number
        sheet.cell(row=next_row, column=13).value = ""  # Additional Phone Number
        sheet.cell(row=next_row, column=14).value = extracted_data.get("jenis_kelamin") or ""  # Gender *
        sheet.cell(row=next_row, column=15).value = extracted_data.get("status_perkawinan") or ""  # Marital Status *
        sheet.cell(row=next_row, column=16).value = extracted_data.get("agama") or ""  # Religion *
        sheet.cell(row=next_row, column=17).value = ""  # Organization Name *
        sheet.cell(row=next_row, column=18).value = ""  # Job Position *
        sheet.cell(row=next_row, column=19).value = ""  # Job Level *
        sheet.cell(row=next_row, column=20).value = ""  # Grade
        sheet.cell(row=next_row, column=21).value = ""  # Class
        sheet.cell(row=next_row, column=22).value = ""  # Employment Status *
        sheet.cell(row=next_row, column=23).value = ""  # Join Date *
        sheet.cell(row=next_row, column=24).value = ""  # End Employment Status Date
        sheet.cell(row=next_row, column=25).value = ""  # Sign Date
        sheet.cell(row=next_row, column=26).value = extracted_data.get("npwp") or ""  # NPWP
        sheet.cell(row=next_row, column=27).value = ""  # Taxable Date
        sheet.cell(row=next_row, column=28).value = extracted_data.get("status_ptkp") or ""  # PTKP Status *
        sheet.cell(row=next_row, column=29).value = ""  # Bank Name
        sheet.cell(row=next_row, column=30).value = extracted_data.get("rekening_bank") or ""  # Bank Account
        sheet.cell(row=next_row, column=31).value = ""  # Bank Account Holder
        sheet.cell(row=next_row, column=32).value = extracted_data.get("bpjs_tk") or ""  # BPJS Ketenagakerjaan
        sheet.cell(row=next_row, column=33).value = extracted_data.get("bpjs_kes") or ""  # BPJS Kesehatan
        sheet.cell(row=next_row, column=34).value = ""  # Resign Date
        sheet.cell(row=next_row, column=35).value = ""  # Branch Name
        sheet.cell(row=next_row, column=36).value = ""  # Type Salary
        sheet.cell(row=next_row, column=37).value = ""  # Overtime Status
        sheet.cell(row=next_row, column=38).value = ""  # Passport
        sheet.cell(row=next_row, column=39).value = ""  # Passport Expired Date
        sheet.cell(row=next_row, column=40).value = extracted_data.get("golongan_darah") or ""  # Blood Type
        sheet.cell(row=next_row, column=41).value = ""  # Postal Code
        sheet.cell(row=next_row, column=42).value = ""  # BPJS Kesehatan Family
        sheet.cell(row=next_row, column=43).value = ""  # Employee Tax Status*
        sheet.cell(row=next_row, column=44).value = ""  # JHT Config
        sheet.cell(row=next_row, column=45).value = ""  # Tax Config
        sheet.cell(row=next_row, column=46).value = ""  # BPJS Kesehatan Config
        sheet.cell(row=next_row, column=47).value = ""  # Jaminan Pensiun Config
        sheet.cell(row=next_row, column=48).value = ""  # NPP BPJS Ketenagakerjaan
        sheet.cell(row=next_row, column=49).value = ""  # Beginning Netto
        sheet.cell(row=next_row, column=50).value = ""  # PPH 21 Paid
        sheet.cell(row=next_row, column=51).value = ""  # Ekspatriat DN Date
        sheet.cell(row=next_row, column=52).value = ""  # Nationality Code
        sheet.cell(row=next_row, column=53).value = ""  # BPJS Ketenagakerjaan Date
        sheet.cell(row=next_row, column=54).value = ""  # BPJS Kesehatan Date
        sheet.cell(row=next_row, column=55).value = ""  # Jaminan Pensiun Date
        sheet.cell(row=next_row, column=56).value = ""  # Payment Schedule
        sheet.cell(row=next_row, column=57).value = ""  # Salary Config
        sheet.cell(row=next_row, column=58).value = ""  # Currency
        sheet.cell(row=next_row, column=59).value = ""  # Cost Center
        sheet.cell(row=next_row, column=60).value = ""  # Cost Center Category
        sheet.cell(row=next_row, column=61).value = ""  # Work Schedule
        sheet.cell(row=next_row, column=62).value = ""  # Overtime Working Day Default
        sheet.cell(row=next_row, column=63).value = ""  # Overtime Day Off Default
        sheet.cell(row=next_row, column=64).value = ""  # Overtime National Holiday Default
        sheet.cell(row=next_row, column=65).value = ""  # Split Payment Policy
        sheet.cell(row=next_row, column=66).value = ""  # Bank Name Secondary
        sheet.cell(row=next_row, column=67).value = ""  # Bank Account  Secondary
        sheet.cell(row=next_row, column=68).value = ""  # Bank Account Holder  Secondary
        sheet.cell(row=next_row, column=69).value = ""  # Bank Name Tertiary
        sheet.cell(row=next_row, column=70).value = ""  # Bank Account Tertiary
        sheet.cell(row=next_row, column=71).value = ""  # Bank Account Holder Tertiary
        sheet.cell(row=next_row, column=72).value = ""  # Jenis Dok. Referensi Bukti Potong
        sheet.cell(row=next_row, column=73).value = ""  # Nomor Dok. Referensi Bukti Potong
        sheet.cell(row=next_row, column=74).value = ""  # Tanggal Dok. Referensi Bukti Potong
        sheet.cell(row=next_row, column=75).value = ""  # TIN (Taxpayer Identification Number)
        
        sheet.cell(row=next_row, column=76).value = extracted_data.get("catatan_dokumen") or ""  # Catatan Dokumen/Error
        
        workbook.save(excel_path)
        print(f"[Success] Data untuk '{employee_meta.get('nama')}' berhasil disisipkan di baris {next_row} (Sheet: Inject).")
        
    except Exception as e:
        print(f"[Error] Gagal menyisipkan data secara spesifik untuk '{employee_meta.get('nama')}': {e}")


def get_processed_employees(excel_path: str) -> tuple[set, set]:
    """
    Membaca sheet 'Inject' untuk mengumpulkan ID dan Nama karyawan yang sudah pernah diproses.
    Berguna untuk fitur 'Resume' atau melewati karyawan yang sudah ada di Excel.
    """
    processed_ids = set()
    processed_names = set()
    
    if not os.path.exists(excel_path):
        return processed_ids, processed_names
        
    try:
        # Gunakan read_only=True untuk mempercepat pembacaan tanpa memuat seluruh format Excel
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        if "Inject" in workbook.sheetnames:
            sheet = workbook["Inject"]
        else:
            sheet = workbook.active
            
        # Loop mulai dari baris 2 (melewati baris header 1)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Kolom 2 (index 1): Employee ID
            emp_id = row[1]
            if emp_id:
                processed_ids.add(str(emp_id).strip())
                
            # Kolom 4 (index 3): First Name, Kolom 5 (index 4): Last Name
            first_name = str(row[3] or "").strip()
            last_name = str(row[4] or "").strip()
            full_name = f"{first_name} {last_name}".strip().lower()
            if full_name:
                processed_names.add(full_name)
                
    except Exception as e:
        print(f"[Warning] Gagal membaca riwayat data Excel untuk proses resume: {e}")
        
    return processed_ids, processed_names
