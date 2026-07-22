import openpyxl

file_path = "data/Template saja.xlsx"
try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print("Sheets available:")
    for sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        sheet = wb[sheet_name]
        headers_row1 = [sheet.cell(row=1, column=i).value for i in range(1, 20)]
        headers_row2 = [sheet.cell(row=2, column=i).value for i in range(1, 20)]
        headers_row3 = [sheet.cell(row=3, column=i).value for i in range(1, 20)]
        print(f"Row 1: {headers_row1}")
        print(f"Row 2: {headers_row2}")
        print(f"Row 3: {headers_row3}")
except Exception as e:
    print(f"Error: {e}")
